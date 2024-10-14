import os
from PyPDF2 import PdfReader
from sentence_transformers import SentenceTransformer
from openpyxl import load_workbook

def load_documents(folder_path='documents'):
    documents = []
    file_names = []
    for filename in os.listdir(folder_path):
        if filename.endswith('.pdf'):
            file_path = os.path.join(folder_path, filename)
            with open(file_path, 'rb') as file:
                pdf = PdfReader(file)
                text = ''
                for page in pdf.pages:
                    text += page.extract_text()
                documents.append(text)
                file_names.append(filename)
    return documents, file_names

def create_embeddings(documents):
    model_name = 'paraphrase-multilingual-MiniLM-L12-v2'
    model_path = os.path.join(os.getcwd(), 'models', model_name)
    if os.path.exists(model_path):
        model = SentenceTransformer(model_path)
    else:
        model = SentenceTransformer(model_name)
        os.makedirs(model_path, exist_ok=True)
        model.save(model_path)
    return model.encode(documents)

def get_documents_and_embeddings():
    documents, file_names = load_documents()
    doc_embeddings = create_embeddings(documents)
    return documents, doc_embeddings, file_names

def load_excel_documents(file_path):
    documents = []
    file_names = []
    wb = load_workbook(file_path, read_only=True)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:  # 첫 번째 열(A열)에 내용이 있는 경우만 처리
            documents.append(row[0])
            file_names.append(f"평가 데이터 - 행 {ws.index(row)}")
    return documents, file_names
